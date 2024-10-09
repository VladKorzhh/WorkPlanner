from tkinter import filedialog as fd
import pandas as pd
from tkinter import messagebox as mb
import datetime as dt
from tkinter import *
import os
import sqlite3
from openpyxl.reader.excel import load_workbook
import customtkinter as ctk


class Base:
    """Класс по работе с базой данных SQLite"""

    PATH = 'E:/Обучение_IT/Projects/PyCharm/Work_Planner/SalesAndServices.sqlite'
    VALUES_LIST = []
    QUOTES = "'"
    TP_LIST = ['Новое ТП', 'Увеличение ММ', 'Изменение точки присоединения', 'Смена собственника',
               'Опосредованное ТП', 'Уменьшение ММ']
    SERVICE_LIST = ['Замена ПУ', 'Замена МПИ', 'Инструментальная проверка', 'Присоединение жил проводов',
                    'Переопломбировка', 'Вывод ПУ из расчетов', 'Осмотр ПУ', 'Прочее (Полухина Т.Н.)']
    APPEAL_LIST = ['Качество', 'Наружное освещение', 'Обследование ЛЭП', 'Расчистка трассы', 'Прочее (Рахманина Е.В.)',
                   'Замена ввода', 'Хищение э/э', 'Обследование ПУ', 'Замена ПУ по жалобе']

    @classmethod
    def insert(cls, login: str, password: str, name: str) -> bool:
        current_name = cls.QUOTES + name + cls.QUOTES
        current_login = cls.QUOTES + login + cls.QUOTES
        current_password = cls.QUOTES + password + cls.QUOTES
        db = sqlite3.connect(cls.PATH)
        cursor = db.cursor()
        cursor.execute(
            f'INSERT INTO Users (ФИО, Логин, Пароль) VALUES ({current_name}, {current_login}, {current_password})')
        db.commit()
        cursor.close()

    @classmethod
    def select(cls):
        db = sqlite3.connect(cls.PATH)
        cursor = db.cursor()
        cursor.execute(
            f'SELECT * FROM Users')
        users = cursor.fetchall()
        return users

    @classmethod
    def get_label_if_date_invalid(cls, value, def_name):
        """Проверят поле ввода даты на ДД.ММ.ГГГГ"""
        if len(cls.VALUES_LIST[value]) > 10 or len(cls.VALUES_LIST[value].split('.')) < 3:
            return ctk.CTkLabel(def_name, text='формат даты должен быть ДД.ММ.ГГГГ', text_color='red',
                                bg_color='#d5d8db').grid(row=value,
                                                         column=3,
                                                         padx=20,
                                                         pady=5,
                                                         sticky=W)
        else:
            return True

    @classmethod
    def get_label_if_priority_invalid(cls, value, def_name):
        """Проверят не пусто ли поле Приоритет"""
        if cls.VALUES_LIST[value] == "" or cls.VALUES_LIST[value] is None:
            return ctk.CTkLabel(def_name, text='Установите приоритет!', text_color='red',
                                bg_color='#d5d8db').grid(row=value,
                                                         column=3,
                                                         padx=20,
                                                         pady=6,
                                                         sticky=W)
        else:
            return True

    @classmethod
    def validation_by_priority(cls, name_table, def_name):
        """Исполняет функцию get_label_if_date_invalid в зависимости от таблицы """
        match name_table:
            case 'Сервисы':
                if cls.get_label_if_priority_invalid(14, def_name):
                    return True
            case 'ТехПрис':
                if cls.get_label_if_priority_invalid(17, def_name):
                    return True
            case _:
                if cls.get_label_if_priority_invalid(12, def_name):
                    return True

    @classmethod
    def validation_by_date(cls, name_table, def_name):
        """Исполняет функцию get_label_if_priority_invalid в зависимости от таблицы """
        match name_table:
            case 'Сервисы':
                if cls.get_label_if_date_invalid(13, def_name):
                    return True
            case 'ТехПрис':
                if cls.get_label_if_date_invalid(16, def_name):
                    return True
            case _:
                if cls.get_label_if_date_invalid(11, def_name):
                    return True

    @classmethod
    def add_row_in_table(cls, name_table, column, values, text, def_name):
        """Добавляет новую строку с данными в базу"""
        for i in values:
            cls.VALUES_LIST.append(i.get())
        cls.VALUES_LIST.append(text.get(1.0, END))
        print(cls.validation_by_priority(name_table, def_name))
        if cls.validation_by_date(name_table, def_name) and cls.validation_by_priority(name_table, def_name):
            db = sqlite3.connect(cls.PATH)
            cursor = db.cursor()
            cursor.execute(f'INSERT INTO {name_table} {tuple(column)} VALUES {tuple(cls.VALUES_LIST)}')
            db.commit()
            cursor.close()
            success = mb.askyesno(message="Задание создано. Продолжить?")
            if success:
                pass
            else:
                def_name.destroy()
        cls.VALUES_LIST = []

    @classmethod
    def route(cls, table, locality, street, *check_var):
        """Объединяет 4 исходные таблицы базы данных в одно представление и передает их в treeview"""
        # Не нашел способа сделать это одним запросом. Разбил на 2 штуки. Так же разбил по условиям счетчики
        # столбцов в конце. Иначе выпадает ошибка Index out of range
        if check_var[0] == 'нет':
            table.delete(*table.get_children())
        locality_for_sql = cls.QUOTES + locality + cls.QUOTES
        street_for_sql = cls.QUOTES + street + cls.QUOTES
        db = sqlite3.connect(cls.PATH)
        cursor = db.cursor()
        if street_for_sql == "''":
            cursor.execute(
                f"""SELECT id, ФИО, [Тип лица], Телефон, [Номер ПУ], [Фазность ПУ], [Населенный пункт], Улица, Дом,
                           Квартира, Задание, [Дата исполнения], Приоритет
                FROM Задания WHERE ([Населенный пункт] = {locality_for_sql}) AND (Исполнено IS NULL)
                UNION ALL SELECT id, ФИО, [Тип лица], Телефон, [Номер ПУ], [Фазность ПУ], [Населенный пункт], Улица, Дом,
                                 Квартира, [Вид услуги] AS Задание, [Дата исполнения], Приоритет
                FROM Сервисы WHERE ([Населенный пункт] = {locality_for_sql}) AND (Исполнено IS NULL)
                UNION ALL SELECT id, ФИО, [Тип лица], Телефон, '' AS [Номер ПУ], '' AS [Фазность ПУ], [Населенный пункт],
                                 Улица, Дом, Квартира, [Причина обращения] AS Задание, [Дата исполнения], Приоритет
                FROM Обращения WHERE ([Населенный пункт] = {locality_for_sql}) AND (Исполнено IS NULL)
                UNION ALL SELECT id, ФИО, [Тип лица], Телефон, '' AS[Номер ПУ], [U,кВ] AS [Фазность ПУ], [Населенный пункт],
                                 Улица, Дом, Квартира, [Тип подключения] AS Задание, [Дата исполнения], Приоритет
                                 FROM ТехПрис WHERE ([Населенный пункт] = {locality_for_sql}) AND (Исполнено IS NULL)""")
        else:
            cursor.execute(
                f"""SELECT id, ФИО, [Тип лица], Телефон, [Номер ПУ], [Фазность ПУ], [Населенный пункт], Улица, Дом,
                           Квартира, Задание, [Дата исполнения], Приоритет
                FROM Задания WHERE ([Населенный пункт] = {locality_for_sql} AND Улица = {street_for_sql}) AND (Исполнено IS NULL)
                UNION ALL SELECT id, ФИО, [Тип лица], Телефон, [Номер ПУ], [Фазность ПУ], [Населенный пункт], Улица, Дом,
                                 Квартира, [Вид услуги] AS Задание, [Дата исполнения], Приоритет
                FROM Сервисы WHERE ([Населенный пункт] = {locality_for_sql} AND Улица = {street_for_sql}) AND (Исполнено IS NULL)
                UNION ALL SELECT id, ФИО, [Тип лица], Телефон, '' AS [Номер ПУ], '' AS [Фазность ПУ], [Населенный пункт],
                                 Улица, Дом, Квартира, [Причина обращения] AS Задание, [Дата исполнения], Приоритет
                FROM Обращения WHERE ([Населенный пункт] = {locality_for_sql} AND Улица = {street_for_sql}) AND (Исполнено IS NULL)
                UNION ALL SELECT id, ФИО, [Тип лица], Телефон, '' AS[Номер ПУ], [U,кВ] AS [Фазность ПУ], [Населенный пункт],
                                 Улица, Дом, Квартира, [Тип подключения] AS Задание, [Дата исполнения], Приоритет
                FROM ТехПрис WHERE ([Населенный пункт] = {locality_for_sql} AND Улица = {street_for_sql}) AND (Исполнено IS NULL)""")
        for i in cursor.fetchall():
            all_columns = (i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7], i[8], i[9], i[10], i[11], i[12])
            cls.VALUES_LIST.append(all_columns)
        if street == '':
            cursor.execute(
                f'SELECT COUNT(id) FROM Сервисы WHERE ([Населенный пункт] = {locality_for_sql}) AND (Исполнено IS NULL)')
            count1 = cursor.fetchone()[0]
            cursor.execute(
                f'SELECT COUNT(id) FROM Задания WHERE ([Населенный пункт] = {locality_for_sql}) AND (Исполнено IS NULL)')
            count2 = cursor.fetchone()[0]
            cursor.execute(
                f'SELECT COUNT(id) FROM Обращения WHERE ([Населенный пункт] = {locality_for_sql}) AND (Исполнено IS NULL)')
            count3 = cursor.fetchone()[0]
            cursor.execute(
                f'SELECT COUNT(id) FROM ТехПрис WHERE ([Населенный пункт] = {locality_for_sql}) AND (Исполнено IS NULL)')
            count4 = cursor.fetchone()[0]
            db.commit()
            cursor.close()
            count = count1 + count2 + count3 + count4
        else:
            cursor.execute(
                f'SELECT COUNT(id) FROM Сервисы WHERE ([Населенный пункт] = {locality_for_sql} AND Улица = {street_for_sql}) AND (Исполнено IS NULL)')
            count1 = cursor.fetchone()[0]
            cursor.execute(
                f'SELECT COUNT(id) FROM Задания WHERE ([Населенный пункт] = {locality_for_sql} AND Улица = {street_for_sql}) AND (Исполнено IS NULL)')
            count2 = cursor.fetchone()[0]
            cursor.execute(
                f'SELECT COUNT(id) FROM Обращения WHERE ([Населенный пункт] = {locality_for_sql} AND Улица = {street_for_sql}) AND (Исполнено IS NULL)')
            count3 = cursor.fetchone()[0]
            cursor.execute(
                f'SELECT COUNT(id) FROM ТехПрис WHERE ([Населенный пункт] = {locality_for_sql} AND Улица = {street_for_sql}) AND (Исполнено IS NULL)')
            count4 = cursor.fetchone()[0]
            db.commit()
            cursor.close()
            count = count1 + count2 + count3 + count4
        for row in range(0, count):
            table.insert('', 'end', values=cls.VALUES_LIST[row])
        cls.VALUES_LIST = []

    @classmethod
    def show_all_raws(cls, table, *check_var):
        if check_var[0] == 'нет':
            table.delete(*table.get_children())
        db = sqlite3.connect(cls.PATH)
        cursor = db.cursor()
        cursor.execute(
            f"""SELECT id, ФИО, [Тип лица], Телефон, [Номер ПУ], [Фазность ПУ], [Населенный пункт], Улица, Дом,
                           Квартира, Задание, [Дата исполнения], Приоритет
                FROM Задания WHERE Исполнено IS NULL
                UNION ALL SELECT id, ФИО, [Тип лица], Телефон, [Номер ПУ], [Фазность ПУ], [Населенный пункт], Улица, Дом,
                                 Квартира, [Вид услуги] AS Задание, [Дата исполнения], Приоритет
                FROM Сервисы WHERE Исполнено IS NULL
                UNION ALL SELECT id, ФИО, [Тип лица], Телефон, '' AS [Номер ПУ], '' AS [Фазность ПУ], [Населенный пункт],
                                 Улица, Дом, Квартира, [Причина обращения] AS Задание, [Дата исполнения], Приоритет
                FROM Обращения WHERE Исполнено IS NULL
                UNION ALL SELECT id, ФИО, [Тип лица], Телефон, '' AS[Номер ПУ], [U,кВ] AS [Фазность ПУ], [Населенный пункт],
                                 Улица, Дом, Квартира, [Тип подключения] AS Задание, [Дата исполнения], Приоритет
                                 FROM ТехПрис WHERE Исполнено IS NULL""")
        for i in cursor.fetchall():
            all_columns = (i[0], i[1], i[2], i[3], i[4], i[5], i[6], i[7], i[8], i[9], i[10], i[11], i[12])
            cls.VALUES_LIST.append(all_columns)
        cursor.execute(
            f'SELECT COUNT(id) FROM Сервисы WHERE Исполнено IS NULL')
        count1 = cursor.fetchone()[0]
        cursor.execute(
            f'SELECT COUNT(id) FROM Задания WHERE Исполнено IS NULL')
        count2 = cursor.fetchone()[0]
        cursor.execute(
            f'SELECT COUNT(id) FROM Обращения WHERE Исполнено IS NULL')
        count3 = cursor.fetchone()[0]
        cursor.execute(
            f'SELECT COUNT(id) FROM ТехПрис WHERE Исполнено IS NULL')
        count4 = cursor.fetchone()[0]
        db.commit()
        cursor.close()
        count = count1 + count2 + count3 + count4
        for row in range(0, count):
            table.insert('', 'end', values=cls.VALUES_LIST[row])
        cls.VALUES_LIST = []

    @classmethod
    def get_table_row(cls, table, parent_window):
        """Обновляет строку базы данных, внося дату исполнения"""
        time_now = dt.datetime.now()
        time_now_str = cls.QUOTES + time_now.strftime("%d.%m.%Y") + cls.QUOTES
        selected = table.selection()
        if selected:
            selected_item = table.selection()[0]
            row_values = table.item(selected_item, option="values")
            db = sqlite3.connect(cls.PATH)
            cursor = db.cursor()
            TASK = True
            for i in cls.TP_LIST:
                if row_values[10] == i:
                    cursor.execute(
                        f'''UPDATE ТехПрис SET Исполнено = {time_now_str} WHERE id = {row_values[0]}''')
                    table.delete(selected_item)
                    db.commit()
                    TASK = False
            for i in cls.SERVICE_LIST:
                if row_values[10] == i:
                    cursor.execute(
                        f'''UPDATE Сервисы SET Исполнено = {time_now_str} WHERE id = {row_values[0]}''')
                    table.delete(selected_item)
                    db.commit()
                    TASK = False
            for i in cls.APPEAL_LIST:
                if row_values[10] == i:
                    cursor.execute(
                        f'''UPDATE Обращения SET Исполнено = {time_now_str} WHERE id = {row_values[0]}''')
                    table.delete(selected_item)
                    db.commit()
                    TASK = False
            if TASK:
                cursor.execute(
                    f'''UPDATE Задания SET Исполнено = {time_now_str} WHERE id = {row_values[0]}''')
                table.delete(selected_item)
                db.commit()
            cursor.close()
            db.close()
        msg = "Задание внесено в список выполненных!"
        mb.showinfo("Информация", msg, parent=parent_window)

    @classmethod
    def upload_route_to_excel(cls, table_columns, table):
        """Выгружает сформированный маршрут из treeview в файл excel на ПК"""
        for i in table.get_children():
            route_row = table.item(i)["values"]
            cls.VALUES_LIST.append(route_row)
        file = fd.asksaveasfilename(filetypes=(("Excel files", "*.xlsx"), ("CSV files", "*.csv")))
        df = pd.DataFrame(cls.VALUES_LIST, columns=table_columns)
        df['Приоритет'] = df['Приоритет'].astype(int)
        df = df.sort_values('Приоритет')
        writer = pd.ExcelWriter(file, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Лист1', index=False)
        for column in df:
            column_width = max(df[column].astype(str).map(len).max(), len(column))
            col_idx = df.columns.get_loc(column)
            writer.sheets['Лист1'].set_column(col_idx, col_idx, column_width)
        writer.close()
        wb = load_workbook(file)
        ws = wb.active
        ws.auto_filter.ref = ws.dimensions
        wb.save(file)
        os.startfile(file)
        cls.VALUES_LIST = []

    @classmethod
    def unload_table(cls, name_table):
        """Выгружает данные соответствующей таблицы в файл excel на ПК"""
        file = fd.asksaveasfilename(
            filetypes=(("Excel files", "*.xlsx"),
                       ("CSV files", "*.csv")))
        conn = sqlite3.connect(cls.PATH)
        df = pd.read_sql(f'SELECT * FROM {name_table}', conn)
        writer = pd.ExcelWriter(file)
        df.to_excel(writer, index=False, sheet_name='Лист1')
        for column in df:
            column_width = max(df[column].astype(str).map(len).max(), len(column))
            col_idx = df.columns.get_loc(column)
            writer.sheets['Лист1'].set_column(col_idx, col_idx, column_width)
        writer.close()
        wb = load_workbook(file)
        ws = wb.active
        ws.auto_filter.ref = ws.dimensions
        wb.save(file)
        os.startfile(file)

    @classmethod
    def make_dump(cls):
        """Создает дамп базы данных"""
        con = sqlite3.connect(cls.PATH)
        file_name = fd.asksaveasfilename(
            filetypes=(("TXT files", "*.txt"),
                       ("SQL files", "*.sql"),
                       ("SQLite files", "*.sqlite")))
        with open(file_name, 'w') as f:
            for line in con.iterdump():
                f.write('%s\n' % line)

    @classmethod
    def sum_service(cls, name):
        db = sqlite3.connect(cls.PATH)
        cursor = db.cursor()
        cursor.execute(f'SELECT SUM(Сумма) FROM {name}')
        nums = cursor.fetchone()
        db.commit()
        cursor.close()
        return str(nums[0]) + ' ' + 'руб.'

    @classmethod
    def market_coverage(cls, name):
        db = sqlite3.connect(cls.PATH)
        cursor = db.cursor()
        cursor.execute(f'SELECT COUNT(id) FROM {name}')
        all_service = cursor.fetchone()
        all_service_int, = all_service
        cursor.execute(f'SELECT COUNT(Приоритет) FROM {name} WHERE Приоритет == "СТП"')
        stp = cursor.fetchone()
        stp_int, = stp
        mc = str((stp_int / all_service_int) * 100)
        mc = mc[0:4]
        return mc + '%'

    @classmethod
    def counter(cls, name_table):
        db = sqlite3.connect(cls.PATH)
        cursor = db.cursor()
        cursor.execute(f'SELECT COUNT(id) FROM {name_table} WHERE "Исполнено" is NULL')
        nums = cursor.fetchone()
        db.commit()
        return int(nums[0])

    @classmethod
    def counter_where(cls, name_table, last_date):
        db = sqlite3.connect(cls.PATH)
        cursor = db.cursor()
        cursor.execute(
            f'SELECT COUNT(id) FROM {name_table} WHERE [Дата исполнения] < {last_date} AND Исполнено is NULL')
        nums_s = cursor.fetchone()
        db.commit()
        return int(nums_s[0])

    @classmethod
    def counter_not_null(cls, name_table, last_date, first_date):
        db = sqlite3.connect(cls.PATH)
        cursor = db.cursor()
        cursor.execute(
            f'SELECT COUNT(id) FROM {name_table} WHERE Исполнено BETWEEN {first_date} AND {last_date}')
        nums = cursor.fetchone()
        db.commit()
        print(first_date)
        print(last_date)
        return int(nums[0])

    @classmethod
    def make_report(cls):
        now = dt.datetime.now()
        last_date = cls.QUOTES + now.strftime("%d.%m.%Y") + cls.QUOTES
        report_date = now - dt.timedelta(days=7)
        first_date = cls.QUOTES + report_date.strftime("%d.%m.%Y") + cls.QUOTES
        all_tusk = cls.counter('Задания')
        all_app = cls.counter('Обращения')
        all_conn = cls.counter('ТехПрис')
        all_serv = cls.counter('Сервисы')
        current_serv = cls.counter_where('Сервисы', last_date)
        current_app = cls.counter_where('Обращения', last_date)
        current_conn = cls.counter_where('ТехПрис', last_date)
        current_tusk = cls.counter_where('Задания', last_date)
        on_date_serv = cls.counter_not_null('Сервисы', last_date, first_date)
        on_date_app = cls.counter_not_null('Обращения', last_date, first_date)
        on_date_conn = cls.counter_not_null('ТехПрис', last_date, first_date)
        on_date_tusk = cls.counter_not_null('Задания', last_date, first_date)
        summ = cls.sum_service('Сервисы')
        mk = cls.market_coverage('Сервисы')
        fn = 'E:/Обучение_IT/Projects/PyCharm/Work_Planner/Форма отчета.xlsx'
        wb = load_workbook(fn)
        ws = wb['Sheet1']
        ws['C16'] = first_date + ' - ' + last_date
        ws['E20'] = all_conn
        ws['E21'] = all_app
        ws['E22'] = all_serv
        ws['E23'] = all_tusk
        ws['F20'] = on_date_conn
        ws['F21'] = on_date_app
        ws['F22'] = on_date_serv
        ws['F23'] = on_date_tusk
        ws['G20'] = current_conn
        ws['G21'] = current_app
        ws['G22'] = current_serv
        ws['G23'] = current_tusk
        ws['C27'] = summ
        ws['C29'] = mk
        wb.save(fn)
        wb.close()
        os.startfile(fn)
